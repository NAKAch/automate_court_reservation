import subprocess
import sys

packages = ['webdriver_manager','openpyxl']
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def upgrade(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package])

# Install/upgrade packages
install('selenium==4.1.0')
for package in packages:
    try:
        dist = __import__(package)
        print("{} ({}) is already installed".format(package, dist.__version__))
        upgrade(package)
        print("{} is upgraded".format(package))
    except ImportError:
        print("{} is NOT installed".format(package))
        install(package)
        print("{} is installed".format(package))


import openpyxl
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

options = Options()

try:
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    print("WebDriver has been successfully initiated.")
except Exception as e:
    print("An error occurred while initiating WebDriver:", e)

wait = WebDriverWait(driver,10)
wait_all_elements = wait.until(EC.presence_of_all_elements_located)

wb = load_workbook('reservation_target_court.xlsx')

def get_filled_sheets_name():
    sheet_names = wb.sheetnames
    filled_sheets_name = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        if not ws['C2'].value == None:
            filled_sheets_name.append(sheet_name)
    return filled_sheets_name

def get_target_court_information(target_sheet):
    user_id = [target_sheet['C2'].value]
    user_password = [target_sheet['C3'].value]
    target_court = [target_sheet['C4'].value]
    user_target_information = [user_id,user_password,target_court]
    user_target_date = []
    number_target_frame = 0
    for row in target_sheet['C8:D17']:
        one_row = []
        for cell in row:
            one_row.append(cell.value)
        if one_row[0] == None or one_row[1] == None:
            pass
        else:
            user_target_date.append(one_row)
            number_target_frame += 1
    user_target_information.append(user_target_date)
    return user_target_information, number_target_frame

def get_all_target_list():
    all_target_list = []
    sum_of_target_frame = 0
    for i in get_filled_sheets_name():
        all_target_list.append(get_target_court_information(wb[i])[0])
        sum_of_target_frame += get_target_court_information(wb[i])[1]
    return all_target_list, sum_of_target_frame

main_list = get_all_target_list()[0]

print(main_list)

import tkinter as tk
from tkinter import messagebox

def web_drive():
    root = tk.Tk()
    root.withdraw()  # tkinterのメインウィンドウを表示しない
    driver.get('https://g-kyoto.pref.kyoto.lg.jp/reserve_j/core_i/init.asp?SBT=1')
    wait_all_elements
    response = messagebox.askokcancel("確認", "抽選コート数は"+str(get_all_target_list()[1])+"個ですか？\n正しく無い場合、「予約するコート」が入力されていない、もしくはエクセルの保存をしていない可能性があります。\n正しい場合は「OK」を選択してください。")
    if response:
        comfirm = messagebox.askokcancel("確認","抽選を実行しますか？\n実行する場合、excelファイルが閉じていることを確認してください。")
        if comfirm:
            messagebox.showinfo("実行","抽選予約を実行します。")
            #予約を実行する
            time.sleep(5)
            driver.switch_to.frame('MainFrame')
            for i in range(len(main_list)):
                date_time_list = main_list[i][3]
                park_name = main_list[i][2][0]
                user_id = main_list[i][0][0]
                user_password = main_list[i][1][0]
                for j in range(len(date_time_list)):
                    time.sleep(5)
                    driver.find_element_by_link_text('テニス').click()
                    wait_all_elements
                    driver.find_element_by_link_text('京都市左京区').click()
                    wait_all_elements
                    driver.find_element_by_name('btn_next').click()
                    wait_all_elements
                    if park_name == '岡崎':
                        driver.find_element_by_xpath('/html/body/form/div[2]/center/table[4]/tbody/tr/td/table/tbody/tr[1]/td[3]/input').click()
                    elif park_name == '宝':
                        driver.find_element_by_xpath('/html/body/form/div[2]/center/table[4]/tbody/tr/td/table/tbody/tr[3]/td[7]/input').click()
                    wait_all_elements
                    driver.find_element_by_xpath('/html/body/form/div[2]/div[1]/center/table[4]/tbody/tr[1]/th[3]/a').click()
                    for one_day_time in date_time_list:
                        driver.find_element_by_link_text(str(one_day_time[0])).click()
                        time_xpath = [driver.find_element_by_xpath('/html/body/form/div[2]/div[2]/left/table[3]/tbody/tr/td/table/tbody/tr[5]/td[1]/a/img'),driver.find_element_by_xpath('/html/body/form/div[2]/div[2]/left/table[3]/tbody/tr/td/table/tbody/tr[5]/td[5]/a/img'),driver.find_element_by_xpath('/html/body/form/div[2]/div[2]/left/table[3]/tbody/tr/td/table/tbody/tr[5]/td[6]/a/img')]
                        if one_day_time[1] == '8~10':
                            time_xpath[0].click()
                        elif one_day_time[1] == '4~6':
                            time_xpath[1].click()
                        elif one_day_time[1] == '6~9':
                            time_xpath[2].click()
                        wait_all_elements

                        original_window = driver.current_window_handle
                        wait_all_elements
                        new_window_handle = driver.window_handles[-1]
                        wait_all_elements
                        driver.switch_to.window(new_window_handle)
                        wait_all_elements
                        # `<select>`要素を特定
                        select_element = Select(driver.find_element_by_tag_name('select'))
                        # オプションを選択。ここでは値（value）によって選択します。
                        wait_all_elements
                        select_element.select_by_value('1')
                        wait_all_elements
                        driver.find_element_by_xpath('/html/body/form/p/input[1]').click()
                        wait_all_elements
                        driver.switch_to.window(original_window)
                        wait_all_elements
                        driver.switch_to.frame('MainFrame')
                        wait_all_elements
                        driver.find_element_by_css_selector('input.clsImage[name="btn_ok"]').click()
                        wait_all_elements

                        try:
                            blank_of_user_id = driver.find_element_by_xpath('/html/body/form/div[2]/center/table/tbody/tr[1]/td/input')
                            blank_of_user_id.send_keys(user_id)
                            driver.find_element_by_xpath('/html/body/form/div[2]/center/table/tbody/tr[2]/td/input').send_keys(user_password)
                            driver.find_element_by_css_selector('input.clsImage[name="btn_ok"]').click()

                        except NoSuchElementException:
                            pass
                        wait_all_elements
                        try:
                            driver.find_element_by_name("btn_next").click()
                        except:
                            messagebox.showinfo("エラー",f"{user_id}さんのコートカードは停止されている、予約制限を超えている、ID・パスワードが間違っている可能性があります。\n手動でログインして原因を確認後、エクセルを修正してください。")
                        wait_all_elements
                        driver.find_element_by_name("btn_next").click()
                        wait_all_elements
                        driver.find_element_by_name("btn_cmd").click()
                        wait_all_elements
                        alert = Alert(driver)
                        alert.accept()
                        wait_all_elements
                        driver.find_element_by_name("btn_back").click()
                        try:
                            driver.find_element_by_class_name('ResultMsg')
                            messagebox.showinfo("エラー",f"{user_id}さんのコートカードではすでに同一コマが予約されている、もしくはカードの有効期限が切れている可能性があります。\n手動でログインして原因を確認後、スプレッドシートを修正してください。")
                        except:
                            pass

                    
                    #抽選予約確認のスクショを撮る
                    wait_all_elements
                    driver.find_element_by_name("btn_MoveMenu").click()
                    wait_all_elements
                    driver.find_element_by_xpath("//*[@value='抽選予約確認']").click()
                    wait_all_elements
                    driver.find_element_by_name("cmdSearch").click()

                    # JavaScriptを使ってページの最下部にスクロールします。
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

                    # スクリーンショットを撮り、保存します。
                    wait_all_elements
                    driver.save_screenshot(f'C:\\Users\\yuton\\抽選予約確認\\screenshot({user_id}).png')
                    time.sleep(3)
                    logout_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.NAME, 'btn_LogOut'))
                    )
                    logout_button.click()

        else:
            messagebox.showinfo("キャンセル","実行がキャンセルされました。")
    else:
        messagebox.showinfo("キャンセル","実行がキャンセルされました。")

web_drive()
