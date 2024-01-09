from openpyxl import load_workbook

wb = load_workbook('reservation_target_court.xlsx')

def delete_fill_in():
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # Clear cells in range C2:C4
        for row in ws['C2:C4']:
            for cell in row:
                cell.value = None
        # Clear cells in range C8:D17
        for row in ws['C8:D17']:
            for cell in row:
                cell.value = None

try:
    delete_fill_in()
    wb.save('reservation_target_court.xlsx')
except Exception:
    pass
